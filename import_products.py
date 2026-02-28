#!/usr/bin/env python3
"""
Import products from Excel file into the POS SQLite database.
- Reads Excel columns: Name_ar/en, Type_ar/en, Price, Barcodes, Category, Brand, Size, ImageLink
- Creates categories from Type_ar
- Inserts all products
- Downloads product images from ImageLink1 URLs
"""

import openpyxl
import sqlite3
import os
import urllib.request
import time
import sys

# Paths
EXCEL_PATH = r'C:\Users\HP\Downloads\Supermarket-Products (2).xlsx'
DB_PATH = r'C:\Users\HP\.gemini\antigravity\scratch\pos-system\build\data\pos_database.db'
IMAGES_DIR = r'C:\Users\HP\.gemini\antigravity\scratch\pos-system\build\data\images'

def main():
    print("=" * 60)
    print("  POS System - Excel Product Importer")
    print("=" * 60)
    
    # Create images directory
    os.makedirs(IMAGES_DIR, exist_ok=True)
    
    # Open Excel
    print(f"\n[1/5] Opening Excel file...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    total_rows = ws.max_row - 1  # exclude header
    print(f"  Found {total_rows} products")
    
    # Connect to SQLite
    print(f"\n[2/5] Connecting to database...")
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Add new columns to Products table if they don't exist
    print(f"\n[3/5] Updating database schema...")
    new_columns = [
        ("name_en", "TEXT DEFAULT ''"),
        ("brand_ar", "TEXT DEFAULT ''"),
        ("brand_en", "TEXT DEFAULT ''"),
        ("size_info", "TEXT DEFAULT ''"),
        ("image_path", "TEXT DEFAULT ''"),
    ]
    for col_name, col_type in new_columns:
        try:
            cursor.execute(f"ALTER TABLE Products ADD COLUMN {col_name} {col_type}")
            print(f"  Added column: {col_name}")
        except sqlite3.OperationalError:
            print(f"  Column already exists: {col_name}")
    conn.commit()
    
    # Extract unique categories (Type_ar) and insert
    print(f"\n[4/5] Creating categories...")
    categories = {}
    # First, get existing categories
    cursor.execute("SELECT id, name FROM Categories")
    for row in cursor.fetchall():
        categories[row[1].strip()] = row[0]
    
    # Scan all types from Excel
    new_cats = set()
    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        type_ar = row[2]  # Type_ar is column 3 (index 2)
        if type_ar and type_ar.strip() and type_ar.strip() not in categories:
            new_cats.add(type_ar.strip())
    
    for cat_name in sorted(new_cats):
        cursor.execute("INSERT INTO Categories (name) VALUES (?)", (cat_name,))
        categories[cat_name] = cursor.lastrowid
        print(f"  Added category: {cat_name} (ID: {categories[cat_name]})")
    conn.commit()
    
    # Insert products
    print(f"\n[5/5] Importing products...")
    inserted = 0
    skipped = 0
    errors = 0
    
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        try:
            name_ar = row[0] if row[0] else ""       # Name_ar
            name_en = row[1] if row[1] else ""       # Name_en
            type_ar = row[2] if row[2] else ""       # Type_ar
            price = row[4] if row[4] else 0          # Price (SAR)
            barcodes = str(row[5]) if row[5] else "" # Barcodes
            brand_ar = row[12] if row[12] else ""    # Brand_ar
            brand_en = row[13] if row[13] else ""    # Brand_en
            size_ar = row[10] if row[10] else ""     # Size_ar
            image_link = row[14] if row[14] else ""  # ImageLink1
            
            if not name_ar:
                skipped += 1
                continue
            
            # Get first barcode (some have multiple separated by comma)
            barcode = barcodes.split(",")[0].strip() if barcodes else ""
            
            # Get category ID
            cat_id = categories.get(type_ar.strip(), 1) if type_ar else 1
            
            # Image filename
            image_filename = ""
            if image_link and image_link != "None" and image_link.startswith("http"):
                # Extract filename from URL
                image_filename = image_link.split("/")[-1]
            
            # Insert product
            cursor.execute("""
                INSERT INTO Products (name, name_en, barcode, category_id, unit_id,
                    buy_price, sell_price, tax_rate, min_quantity, current_stock,
                    brand_ar, brand_en, size_info, image_path, active)
                VALUES (?, ?, ?, ?, 1, 0, ?, 15, 0, 0, ?, ?, ?, ?, 1)
            """, (
                name_ar.strip(),
                name_en.strip() if name_en else "",
                barcode if barcode else None,
                cat_id,
                float(price) if price else 0,
                brand_ar.strip() if brand_ar else "",
                brand_en.strip() if brand_en else "",
                size_ar.strip() if size_ar else "",
                image_filename,
            ))
            inserted += 1
            
            if i % 500 == 0:
                conn.commit()
                print(f"  Progress: {i}/{total_rows} ({i*100//total_rows}%)")
                
        except Exception as e:
            errors += 1
            if errors <= 5:
                print(f"  Error row {i}: {e}")
    
    conn.commit()
    print(f"\n  Done! Inserted: {inserted}, Skipped: {skipped}, Errors: {errors}")
    
    # Download images
    print(f"\n[BONUS] Downloading product images...")
    downloaded = 0
    img_errors = 0
    
    # Re-read Excel for image links
    wb2 = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws2 = wb2.active
    
    for i, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), 1):
        image_link = row[14] if row[14] else ""  # ImageLink1
        
        if not image_link or image_link == "None" or not image_link.startswith("http"):
            continue
            
        filename = image_link.split("/")[-1]
        filepath = os.path.join(IMAGES_DIR, filename)
        
        if os.path.exists(filepath):
            continue  # Already downloaded
        
        try:
            urllib.request.urlretrieve(image_link, filepath)
            downloaded += 1
            if downloaded % 100 == 0:
                print(f"  Downloaded {downloaded} images...")
        except Exception as e:
            img_errors += 1
            if img_errors <= 3:
                print(f"  Image download error: {e}")
    
    wb2.close()
    print(f"  Images downloaded: {downloaded}, Errors: {img_errors}")
    
    # Summary
    cursor.execute("SELECT COUNT(*) FROM Products WHERE active = 1")
    total = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(DISTINCT category_id) FROM Products WHERE active = 1")
    cat_count = cursor.fetchone()[0]
    
    print(f"\n{'=' * 60}")
    print(f"  IMPORT COMPLETE!")
    print(f"  Total products in DB: {total}")
    print(f"  Categories used: {cat_count}")
    print(f"  Images folder: {IMAGES_DIR}")
    print(f"{'=' * 60}")
    
    conn.close()
    wb.close()

if __name__ == "__main__":
    main()

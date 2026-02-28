import openpyxl
import json

wb = openpyxl.load_workbook(r'C:\Users\HP\Downloads\Supermarket-Products (2).xlsx')
ws = wb.active
print(f"Sheet: {ws.title}")
print(f"Rows: {ws.max_row}")
print(f"Cols: {ws.max_column}")
print()

print("=== HEADERS ===")
headers = []
for i, cell in enumerate(ws[1], 1):
    headers.append(cell.value)
    print(f"  Col {i}: {cell.value}")

print()
print("=== FIRST 3 DATA ROWS ===")
for r in range(2, min(5, ws.max_row + 1)):
    print(f"\n--- Row {r} ---")
    for i, cell in enumerate(ws[r], 1):
        print(f"  {headers[i-1]}: {cell.value}")
